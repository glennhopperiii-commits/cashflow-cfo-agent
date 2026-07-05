"""The sponsor deliverable: a traditional 13-week cash flow workbook.

Builds the classic FP&A treasury layout — line items down the side, weeks
across the top — from the same data and engine the agent used, then writes
it as a formatted Excel workbook (always) or a Google Sheet (from Colab,
after google.colab.auth). Every line ties to the engine's weekly totals to
the dollar; build_forecast_grid() asserts it.

Row model shared by both writers:
    {"label": str, "values": [13 floats] | None, "total": float | None,
     "style": "title" | "note" | "header" | "item" | "total" | "metric"
              | "flag" | "spacer"}
"""

from collections import defaultdict
from datetime import date, timedelta

from . import engine, tools
from .config import OUTPUT_DIR

TOP_CUSTOMER_ROWS = 5


def _customer_collections(invoices, week1, num_weeks):
    """AR collections bucketed by customer x week (in-horizon only)."""
    by_cust = defaultdict(lambda: [0.0] * num_weeks)
    totals = defaultdict(float)
    for inv in invoices:
        totals[inv["customer"]] += inv["amount"]
        w = engine.date_to_week(engine.expected_pay_date(inv), week1, num_weeks)
        if w <= num_weeks:
            by_cust[inv["customer"]][w - 1] += inv["amount"]
    ranked = sorted(totals, key=lambda c: -totals[c])
    return by_cust, ranked


def build_forecast_grid() -> dict:
    facility = tools._load_facility()
    invoices = tools._load_invoices()
    bills = tools._load_bills()
    fixed = tools._load_fixed()
    sales = tools._load_sales()
    week1 = tools._week1(facility)
    num_weeks = facility["num_weeks"]

    det = tools._run_deterministic(facility, invoices, bills, fixed, sales)
    weeks = det["weeks"]

    # --- Receipts: top customers, all other, forecast new billings ---------
    by_cust, ranked = _customer_collections(invoices, week1, num_weeks)
    top = ranked[:TOP_CUSTOMER_ROWS]
    other = [0.0] * num_weeks
    for c in ranked[TOP_CUSTOMER_ROWS:]:
        other = [a + b for a, b in zip(other, by_cust[c])]

    new_billings = [0.0] * num_weeks
    for row in sales:
        for amount, lag, _src in engine.split_sales_row(row):
            collect = date.fromisoformat(str(row["week_start"])) + timedelta(days=int(lag))
            w = engine.date_to_week(collect, week1, num_weeks)
            if w <= num_weeks:
                new_billings[w - 1] += amount

    receipt_rows = [
        {"label": f"{c}", "values": by_cust[c]} for c in top
    ] + [
        {"label": "All other customers", "values": other},
        {"label": "Collections on forecast new billings", "values": new_billings},
    ]
    total_receipts = [sum(r["values"][i] for r in receipt_rows) for i in range(num_weeks)]

    # --- Disbursements: trade AP + fixed schedule by category --------------
    ap_weekly = [0.0] * num_weeks
    for b in bills:
        w = engine.date_to_week(date.fromisoformat(b["due_date"]), week1, num_weeks)
        if w <= num_weeks:
            ap_weekly[w - 1] += b["amount"]

    cat_weekly = defaultdict(lambda: [0.0] * num_weeks)
    for r in fixed:
        if 1 <= r["week"] <= num_weeks:
            cat_weekly[r["category"]][r["week"] - 1] += r["amount"]

    category_labels = [
        ("materials forecast", "Material purchases (forecast ramp POs)"),
        ("payroll", "Payroll (biweekly)"),
        ("benefits", "Health benefits"),
        ("occupancy", "Rent"),
        ("operations", "Utilities & plant services"),
        ("insurance", "Insurance premium (semiannual)"),
        ("tax", "Estimated income tax"),
        ("capex", "Capex deposit (5-axis machining center)"),
        ("acquisition", "Seller note earn-out"),
        ("debt service", "Term loan debt service"),
    ]
    disb_rows = [{"label": "Trade AP (open bills)", "values": ap_weekly}] + [
        {"label": label, "values": cat_weekly[cat]}
        for cat, label in category_labels if cat in cat_weekly
    ]
    total_disb = [sum(r["values"][i] for r in disb_rows) for i in range(num_weeks)]

    # --- Tie out to the engine, to the dollar ------------------------------
    for i, w in enumerate(weeks):
        assert abs(total_receipts[i] - w["receipts"]) < 1, f"receipts mismatch wk {i + 1}"
        assert abs(total_disb[i] - w["disbursements"]) < 1, f"disbursements mismatch wk {i + 1}"

    # --- Assemble the render rows ------------------------------------------
    week_endings = [week1 + timedelta(days=4 + 7 * i) for i in range(num_weeks)]
    covenant = facility["covenant"]

    def row(label, values=None, style="item", total="sum"):
        return {
            "label": label,
            "values": list(values) if values is not None else None,
            "total": (sum(values) if (values is not None and total == "sum")
                      else (total if total != "sum" else None)),
            "style": style,
        }

    beginning = [facility["beginning_cash"]] + [w["ending_cash"] for w in weeks[:-1]]

    rows = [
        row(facility["company"], style="title"),
        row("13-Week Cash Flow Forecast — prepared for Granite Peak Partners", style="note"),
        row(f"As of the {facility['as_of_close']} close · $ whole dollars · "
            f"covenant: cash + availability ≥ ${covenant['threshold']:,.0f} at week {covenant['test_week']}",
            style="note"),
        row("", style="spacer"),
        row("CASH RECEIPTS", style="header"),
        *[row(r["label"], r["values"]) for r in receipt_rows],
        row("Total cash receipts", total_receipts, style="total"),
        row("", style="spacer"),
        row("CASH DISBURSEMENTS", style="header"),
        *[row(r["label"], r["values"]) for r in disb_rows],
        row("Total cash disbursements", total_disb, style="total"),
        row("", style="spacer"),
        row("CASH POSITION", style="header"),
        row("Beginning cash", beginning, total=None),
        row("Net cash flow", [w["net_flow"] for w in weeks], style="total"),
        row("Cash before revolver", [w["pre_revolver_cash"] for w in weeks], total=None),
        row("Revolver draw", [w["revolver_draw"] for w in weeks]),
        row("Revolver paydown", [-w["revolver_paydown"] for w in weeks]),
        row("Ending cash", [w["ending_cash"] for w in weeks], style="total", total=None),
        row(f"Memo: operating cash floor ${facility['operating_cash_floor']:,.0f}", style="note"),
        row("", style="spacer"),
        row("REVOLVER & COVENANT", style="header"),
        row("Revolver balance", [w["revolver_balance"] for w in weeks], style="metric", total=None),
        row("Undrawn availability", [w["availability"] for w in weeks], style="metric", total=None),
        row("Covenant liquidity (cash + availability)",
            [w["covenant_liquidity"] for w in weeks], style="total", total=None),
        row("Covenant threshold", [covenant["threshold"]] * num_weeks, style="metric", total=None),
        row("Headroom vs covenant",
            [w["covenant_liquidity"] - covenant["threshold"] for w in weeks],
            style="total", total=None),
        row(f"Week-{covenant['test_week']} covenant test: "
            f"{'PASS' if det['covenant_pass'] else 'BREACH'} "
            f"(headroom ${det['covenant_headroom']:,.0f})", style="flag"),
    ]

    return {
        "rows": rows,
        "week_labels": [f"Week {i + 1}" for i in range(num_weeks)],
        "week_endings": [d.isoformat() for d in week_endings],
        "num_weeks": num_weeks,
        "covenant_pass": det["covenant_pass"],
        "trough_week": det["trough_week"],
        "trough_cash": det["trough_cash"],
    }


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_xlsx(grid: dict | None = None, path=None) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    grid = grid or build_forecast_grid()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = path or OUTPUT_DIR / "cashflow_13wk.xlsx"

    INK, SLATE, LIGHT, RED, GREEN = "1E293B", "3B5998", "F1F5F9", "DC2626", "16A34A"
    NUMFMT = "#,##0;(#,##0);—"
    n = grid["num_weeks"]

    wb = Workbook()
    ws = wb.active
    ws.title = "13-Week Cash Flow"
    ws.sheet_view.showGridLines = False

    header_row_ix = None
    r = 1
    for row in grid["rows"]:
        style = row["style"]
        cell = ws.cell(row=r, column=1, value=row["label"] or None)

        if style == "title":
            cell.font = Font(size=15, bold=True, color=INK)
        elif style == "note":
            cell.font = Font(size=10, italic=True, color="64748B")
        elif style == "header":
            if header_row_ix is None:
                # First section header: put the week header line right above it
                ws.insert_rows(r)
                ws.cell(row=r, column=1, value="Line item").font = Font(bold=True, color="FFFFFF")
                for i in range(n):
                    c = ws.cell(row=r, column=2 + i,
                                value=f"{grid['week_labels'][i]}\n{grid['week_endings'][i][5:]}")
                    c.font = Font(bold=True, color="FFFFFF", size=10)
                    c.alignment = Alignment(horizontal="right", wrap_text=True)
                for col in range(1, n + 3):
                    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=SLATE)
                ws.cell(row=r, column=n + 2, value="Total").font = Font(bold=True, color="FFFFFF")
                ws.cell(row=r, column=n + 2).alignment = Alignment(horizontal="right")
                header_row_ix = r
                r += 1
                cell = ws.cell(row=r, column=1, value=row["label"])
            cell.font = Font(bold=True, color=SLATE, size=11)
            for col in range(1, n + 3):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)
        elif style in ("item", "metric"):
            cell.font = Font(color=INK, size=11)
            cell.alignment = Alignment(indent=1)
        elif style == "total":
            cell.font = Font(bold=True, color=INK, size=11)
        elif style == "flag":
            cell.font = Font(bold=True, size=11,
                             color=(GREEN if grid["covenant_pass"] else RED))

        if row["values"] is not None:
            for i, v in enumerate(row["values"]):
                c = ws.cell(row=r, column=2 + i, value=round(v))
                c.number_format = NUMFMT
                c.font = Font(bold=(style == "total"), color=INK, size=11)
                if style == "total":
                    c.border = Border(top=Side(style="thin", color=SLATE))
            if row["total"] is not None:
                c = ws.cell(row=r, column=n + 2, value=round(row["total"]))
                c.number_format = NUMFMT
                c.font = Font(bold=True, color=INK, size=11)
        r += 1

    ws.column_dimensions["A"].width = 40
    for i in range(n + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 12
    if header_row_ix:
        ws.freeze_panes = ws.cell(row=header_row_ix + 1, column=2)

    wb.save(path)
    return str(path)


# ---------------------------------------------------------------------------
# Google Sheets writer (Colab-friendly)
# ---------------------------------------------------------------------------

def to_google_sheet(grid: dict | None = None, title: str = "Cascade Precision — 13-Week Cash Flow") -> str:
    """Create a Google Sheet and return its URL.

    In Colab: run `from google.colab import auth; auth.authenticate_user()`
    first, then this uses your Google account directly (gspread is
    preinstalled). Locally: needs gspread plus application-default
    credentials with the Sheets scope.
    """
    try:
        import google.auth
        import gspread
    except ImportError as exc:
        raise RuntimeError(
            "gspread/google-auth not available. In Colab run "
            "`from google.colab import auth; auth.authenticate_user()` first; "
            "locally `pip install gspread google-auth`."
        ) from exc

    creds, _ = google.auth.default(scopes=[
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
    ])
    gc = gspread.authorize(creds)

    grid = grid or build_forecast_grid()
    n = grid["num_weeks"]

    values, formats = [], []  # formats: (row_ix_1based, style)
    week_header = ["Line item"] + [
        f"{grid['week_labels'][i]} ({grid['week_endings'][i][5:]})" for i in range(n)
    ] + ["Total"]

    header_inserted = False
    for row in grid["rows"]:
        if row["style"] == "header" and not header_inserted:
            values.append(week_header)
            formats.append((len(values), "colhead"))
            header_inserted = True
        line = [row["label"]]
        if row["values"] is not None:
            line += [round(v) for v in row["values"]]
            line += [round(row["total"])] if row["total"] is not None else [""]
        values.append(line)
        formats.append((len(values), row["style"]))

    sh = gc.create(title)
    ws = sh.sheet1
    ws.update(values, value_input_option="RAW")

    slate, light = {"red": 0.23, "green": 0.35, "blue": 0.60}, {"red": 0.95, "green": 0.96, "blue": 0.98}
    requests = [
        {"updateSheetProperties": {"properties": {
            "sheetId": ws.id, "gridProperties": {"frozenRowCount": 0, "frozenColumnCount": 1}},
            "fields": "gridProperties.frozenColumnCount"}},
        {"repeatCell": {"range": {"sheetId": ws.id, "startColumnIndex": 1, "endColumnIndex": n + 2},
                        "cell": {"userEnteredFormat": {"numberFormat": {
                            "type": "NUMBER", "pattern": "#,##0;(#,##0);—"}}},
                        "fields": "userEnteredFormat.numberFormat"}},
        {"updateDimensionProperties": {"range": {"sheetId": ws.id, "dimension": "COLUMNS",
                                                 "startIndex": 0, "endIndex": 1},
                                       "properties": {"pixelSize": 300}, "fields": "pixelSize"}},
    ]
    for r_ix, style in formats:
        rng = {"sheetId": ws.id, "startRowIndex": r_ix - 1, "endRowIndex": r_ix}
        if style == "colhead":
            requests.append({"repeatCell": {"range": rng, "cell": {"userEnteredFormat": {
                "backgroundColor": slate,
                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}}}},
                "fields": "userEnteredFormat(backgroundColor,textFormat)"}})
        elif style == "header":
            requests.append({"repeatCell": {"range": rng, "cell": {"userEnteredFormat": {
                "backgroundColor": light, "textFormat": {"bold": True}}},
                "fields": "userEnteredFormat(backgroundColor,textFormat)"}})
        elif style in ("total", "title", "flag"):
            requests.append({"repeatCell": {"range": rng, "cell": {"userEnteredFormat": {
                "textFormat": {"bold": True}}}, "fields": "userEnteredFormat.textFormat"}})
    sh.batch_update({"requests": requests})
    return sh.url
